# extractors/xlsx_extractor.py
"""
XLSX extractor using openpyxl.

Reads every sheet and concatenates calculated cell values and comments into a
single string.
"""

import os
from openpyxl import load_workbook

from .errors import ExtractionError


def extract_xlsx(filepath: str) -> str:
    """
    Read cell values and comments from all sheets and return combined text.

    Raises ExtractionError for:
    - None input
    - Non-string types
    - Empty strings
    - Invalid file paths
    """
    # Validate input type.
    if not isinstance(filepath, str):
        raise ExtractionError(f"TypeError: filepath must be str, got {type(filepath).__name__}")

    # Validate non-empty string.
    if not filepath or not filepath.strip():
        raise ExtractionError("ValueError: filepath cannot be empty")

    # Verify that the file exists before opening it.
    if not os.path.isfile(filepath):
        raise ExtractionError("FileNotFoundError: file does not exist")

    try:
        wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    except Exception as exc:
        raise ExtractionError.from_exception(exc) from exc

    parts = []
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_vals = [
                    str(c).strip()
                    for c in row
                    if c is not None and str(c).strip() != ""
                ]
                if row_vals:
                    parts.append(" ".join(row_vals))
    except Exception as exc:
        raise ExtractionError.from_exception(exc) from exc

    # Cell comments aren't exposed by read_only mode (ReadOnlyCell has no
    # .comment), so they'd otherwise be silently dropped even though they can
    # carry PII. Perform a second fully loaded pass only for comments while
    # retaining read_only=True memory savings for the value pass.
    try:
        wb_full = load_workbook(filename=filepath, read_only=False, data_only=True)
        for sheet in wb_full.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    comment = getattr(cell, "comment", None)
                    if comment is not None and comment.text and comment.text.strip():
                        parts.append(comment.text.strip())
    except Exception:
        pass

    return "\n".join(parts)


if __name__ == "__main__":
    print(extract_xlsx("tests/sample_data/example.xlsx") if False else "No test file")
